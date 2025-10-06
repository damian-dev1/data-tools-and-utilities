import os
import re
import csv
import time
import threading
from typing import List, Dict, Optional, Tuple
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import pandas as pd
import openpyxl

# Optional ttkbootstrap theme support (non-breaking fallback to ttk)
try:
    from ttkbootstrap import Style
    _HAS_TTKB = True
except Exception:
    Style = None
    _HAS_TTKB = False


def to_snake_case(header: str) -> str:
    header = re.sub(r'[^a-zA-Z0-9]+', '_', str(header))
    return header.strip('_').lower()


class CancelToken:
    def __init__(self) -> None:
        self._flag = False

    def cancel(self) -> None:
        self._flag = True

    def is_cancelled(self) -> bool:
        return self._flag


class ExcelToCSVApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Excel → CSV Converter")
        self.root.minsize(950, 560)
        try:
            self.root.iconbitmap(default="")
        except Exception:
            pass

        # Optional dark theme
        if _HAS_TTKB:
            self.style = Style(theme="darkly")
        else:
            self.style = None

        # State
        self.file_path: Optional[str] = None
        self.sheet_names: List[str] = []
        self.headers: List[str] = []
        self.header_vars: Dict[str, tk.BooleanVar] = {}
        self._worker_thread: Optional[threading.Thread] = None
        self._cancel = CancelToken()

        self.preview_rows_var = tk.IntVar(value=10)
        self.status_var = tk.StringVar(value="Ready")

        # ---------- Top Bar ----------
        self._build_topbar()

        # ---------- Middle (Left Sidebar + Right Content) ----------
        self.center = ttk.Panedwindow(self.root, orient="horizontal")
        self.center.pack(fill="both", expand=True)

        self.left = ttk.Frame(self.center, padding=(10, 8))
        self.right = ttk.Frame(self.center, padding=(8, 8))
        self.center.add(self.left, weight=0)   # sidebar
        self.center.add(self.right, weight=3)  # main

        self._build_left()
        self._build_right()
        self._bind_context_menus()

        # ---------- Bottom Bar ----------
        self._build_bottombar()

    # ----------------------------- UI: Top & Bottom Bars -----------------------------
    def _build_topbar(self):
        top = ttk.Frame(self.root, padding=(10, 6))
        top.pack(side="top", fill="x")

        title = ttk.Label(top, text="Excel → CSV Converter", font=("Segoe UI", 13, "bold"))
        title.pack(side="left")

        top_spacer = ttk.Label(top, text="")  # flex
        top_spacer.pack(side="left", expand=True, fill="x")

        # Quick actions on the right (non-destructive; duplicates exist in sidebar Actions)
        ttk.Button(top, text="Open Excel…", command=self.select_file).pack(side="right", padx=(6, 0))
        ttk.Button(top, text="Export CSV", command=self.export_csv_fast).pack(side="right", padx=(6, 0))
        ttk.Button(top, text="Preview", command=self.preview_data).pack(side="right")

    def _build_bottombar(self):
        bottom = ttk.Frame(self.root, padding=(10, 6))
        bottom.pack(side="bottom", fill="x")

        # Status (left)
        self.status_lbl = ttk.Label(bottom, textvariable=self.status_var)
        self.status_lbl.pack(side="left")

        # Spacer
        ttk.Label(bottom, text="").pack(side="left", expand=True, fill="x")

        # Progress + Cancel (right)
        self.prog = ttk.Progressbar(bottom, mode="determinate", maximum=100, length=260)
        self.prog.pack(side="right")
        self.cancel_btn = ttk.Button(bottom, text="Cancel", command=self.cancel_current, width=10)
        self.cancel_btn.pack(side="right", padx=(0, 8))
        self.cancel_btn.configure(state="disabled")

    # ----------------------------- UI: Left Sidebar (Logical Groups) -----------------------------
    def _build_left(self):
        # Grid expansion
        self.left.columnconfigure(0, weight=1)

        # 1) Source
        src = ttk.LabelFrame(self.left, text="1. Source", padding=(10, 8))
        src.grid(row=0, column=0, sticky="nsew", pady=(0, 8))
        ttk.Button(src, text="Select Excel File", command=self.select_file).grid(row=0, column=0, sticky="w")
        self.file_lbl = ttk.Label(src, text="", foreground="#888")
        self.file_lbl.grid(row=0, column=1, sticky="w", padx=(8, 0))
        src.columnconfigure(1, weight=1)

        # 2) Sheet & Header
        sh = ttk.LabelFrame(self.left, text="2. Sheet & Header", padding=(10, 8))
        sh.grid(row=1, column=0, sticky="nsew", pady=(0, 8))
        ttk.Label(sh, text="Sheet:").grid(row=0, column=0, sticky="w")
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(sh, textvariable=self.sheet_var, state="readonly")
        self.sheet_combo.grid(row=0, column=1, sticky="we", padx=(6, 0))
        self.sheet_combo.bind("<<ComboboxSelected>>", lambda e: self.reset_headers())
        ttk.Label(sh, text="Header Row:").grid(row=1, column=0, sticky="w", pady=(6, 0))
        self.header_row_var = tk.StringVar(value="1")
        self.header_row_spin = ttk.Spinbox(sh, from_=1, to=9999, textvariable=self.header_row_var, width=8)
        self.header_row_spin.grid(row=1, column=1, sticky="w", padx=(6, 0), pady=(6, 0))
        ttk.Button(sh, text="Load Headers", command=self.load_headers).grid(row=2, column=0, columnspan=2, sticky="e", pady=(8, 0))
        sh.columnconfigure(1, weight=1)

        # 3) Options
        opts = ttk.LabelFrame(self.left, text="3. Options", padding=(10, 8))
        opts.grid(row=2, column=0, sticky="nsew", pady=(0, 8))
        self.opt_snake = tk.BooleanVar(value=True)
        self.opt_dedup = tk.BooleanVar(value=True)
        ttk.Checkbutton(opts, text="snake_case headers", variable=self.opt_snake).grid(row=0, column=0, sticky="w")
        ttk.Checkbutton(opts, text="Remove duplicate rows", variable=self.opt_dedup).grid(row=1, column=0, sticky="w")

        # 4) Columns (scrollable)
        cols = ttk.LabelFrame(self.left, text="4. Columns", padding=(10, 8))
        cols.grid(row=3, column=0, sticky="nsew", pady=(0, 8))
        cols.rowconfigure(0, weight=1)
        cols.columnconfigure(0, weight=1)

        self.columns_frame = ttk.Frame(cols)
        self.columns_frame.grid(row=0, column=0, sticky="nsew")

        self.col_canvas = tk.Canvas(self.columns_frame, borderwidth=0, highlightthickness=0)
        self.col_scroll_y = ttk.Scrollbar(self.columns_frame, orient="vertical", command=self.col_canvas.yview)
        self.col_list = ttk.Frame(self.col_canvas)
        self.col_list_id = self.col_canvas.create_window((0, 0), window=self.col_list, anchor="nw")

        self.col_canvas.configure(yscrollcommand=self.col_scroll_y.set)
        self.col_canvas.pack(side="left", fill="both", expand=True)
        self.col_scroll_y.pack(side="right", fill="y")
        self.col_list.bind("<Configure>", lambda e: self.col_canvas.configure(scrollregion=self.col_canvas.bbox("all")))
        self.col_canvas.bind("<Configure>", lambda e: self.col_canvas.itemconfigure(self.col_list_id, width=e.width))

        # Select/Deselect buttons
        selrow = ttk.Frame(cols)
        selrow.grid(row=1, column=0, sticky="we", pady=(8, 0))
        ttk.Button(selrow, text="Select All", command=lambda: self.set_all_checkboxes(True)).pack(side="left")
        ttk.Button(selrow, text="Deselect All", command=lambda: self.set_all_checkboxes(False)).pack(side="left", padx=(8, 0))

        # 5) Actions
        act = ttk.LabelFrame(self.left, text="5. Actions", padding=(10, 8))
        act.grid(row=4, column=0, sticky="we")
        ttk.Label(act, text="Preview rows:").pack(side="left")
        ttk.Spinbox(act, from_=1, to=5000, textvariable=self.preview_rows_var, width=6).pack(side="left", padx=(6, 8))
        ttk.Button(act, text="Preview", command=self.preview_data).pack(side="left")
        ttk.Button(act, text="Export CSV", command=self.export_csv_fast).pack(side="left", padx=(8, 0))

        # Let Columns frame expand vertically
        self.left.rowconfigure(3, weight=1)

    # ----------------------------- UI: Right (Preview + Log) -----------------------------
    def _build_right(self):
        # Preview header
        hdr = ttk.Frame(self.right)
        hdr.pack(fill="x")
        ttk.Label(hdr, text="Data Preview", font=("Segoe UI", 12, "bold")).pack(side="left")

        # Tree (preview grid)
        tv_wrap = ttk.Frame(self.right)
        tv_wrap.pack(fill="both", expand=True, pady=(8, 6))
        tv_wrap.rowconfigure(0, weight=1)
        tv_wrap.columnconfigure(0, weight=1)

        self.tree = ttk.Treeview(tv_wrap, columns=(), show="headings")
        yscroll = ttk.Scrollbar(tv_wrap, orient="vertical", command=self.tree.yview)
        xscroll = ttk.Scrollbar(tv_wrap, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="we")

        self.tree.bind("<Button-3>", self._tree_context_menu)
        self.tree.bind("<Control-c>", lambda e: self._copy_selected_rows())
        self.tree.bind("<Control-C>", lambda e: self._copy_selected_rows())

        # Logs
        log_frame = ttk.LabelFrame(self.right, text="Log")
        log_frame.pack(fill="x", pady=(0, 4))
        self.log_text = tk.Text(log_frame, height=6, wrap="word")
        self.log_text.pack(fill="both", expand=True)

    # ----------------------------- Context Menus -----------------------------
    def _bind_context_menus(self):
        self._entry_menu = tk.Menu(self.root, tearoff=0)
        self._entry_menu.add_command(label="Paste", command=lambda: self.root.focus_get().event_generate("<<Paste>>"))
        self._entry_menu.add_command(label="Copy", command=lambda: self.root.focus_get().event_generate("<<Copy>>"))
        self._entry_menu.add_command(label="Cut", command=lambda: self.root.focus_get().event_generate("<<Cut>>"))
        self._entry_menu.add_separator()

        for widget in (self.sheet_combo, self.header_row_spin,):
            widget.bind("<Button-3>", self._show_entry_menu)
        self.root.bind_class("TEntry", "<Button-3>", self._show_entry_menu)
        self.root.bind_class("TCombobox", "<Button-3>", self._show_entry_menu)
        self.root.bind_class("TSpinbox", "<Button-3>", self._show_entry_menu)
        self.root.bind_class("Text", "<Button-3>", self._show_entry_menu)
        self.root.bind_class("Treeview", "<Button-3>", self._tree_context_menu)

        self._tree_menu = tk.Menu(self.root, tearoff=0)
        self._tree_menu.add_command(label="Copy selected row(s)", command=self._copy_selected_rows)

    def _show_entry_menu(self, event):
        try:
            self._entry_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self._entry_menu.grab_release()

    def _tree_context_menu(self, event):
        try:
            item = self.tree.identify_row(event.y)
            if item:
                self.tree.selection_set(item)
            self._tree_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self._tree_menu.grab_release()

    # ----------------------------- Helpers: Grid/Status/Log -----------------------------
    def _copy_selected_rows(self):
        sel = self.tree.selection()
        if not sel:
            return
        rows = []
        cols = self.tree["columns"]
        for iid in sel:
            values = self.tree.item(iid, "values")
            row_dict = {c: v for c, v in zip(cols, values)}
            rows.append(row_dict)
        lines = ["\t".join(cols)]
        for rd in rows:
            lines.append("\t".join(str(rd.get(c, "")) for c in cols))
        txt = "\n".join(lines)
        self.root.clipboard_clear()
        self.root.clipboard_append(txt)

    def set_status(self, text: str, ok: bool = True):
        self.status_var.set(text)
        # green for ok, red for error (fallback color on non-ttkbootstrap still fine)
        self.status_lbl.configure(foreground="#2e7d32" if ok else "#c62828")
        self._log(text)

    def _log(self, msg: str):
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")

    def reset_headers(self):
        self.headers = []
        self.header_vars.clear()
        for w in list(self.col_list.children.values()):
            w.destroy()
        self.clear_preview()
        self.set_status("")

    def set_all_checkboxes(self, value: bool):
        for v in self.header_vars.values():
            v.set(value)

    def get_selected_columns(self) -> List[str]:
        return [c for c, v in self.header_vars.items() if v.get()]

    def clear_preview(self):
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = ()

    def render_preview(self, df: pd.DataFrame):
        self.clear_preview()
        cols = list(df.columns)
        self.tree["columns"] = cols
        for c in cols:
            self.tree.heading(c, text=str(c))
            self.tree.column(c, width=140, stretch=True)
        for _, row in df.iterrows():
            values = [("" if pd.isna(v) else v) for v in row.tolist()]
            self.tree.insert("", "end", values=values)

    # ----------------------------- File / Data Ops -----------------------------
    def select_file(self):
        path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm *.xls"), ("All files", "*.*")]
        )
        if not path:
            return
        self.file_path = path
        self.file_lbl.configure(text=os.path.basename(path))
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            self.sheet_names = wb.sheetnames
            self.sheet_combo["values"] = self.sheet_names
            if self.sheet_names:
                self.sheet_var.set(self.sheet_names[0])
            self.set_status("File loaded. Pick sheet & header row, then Load Headers.", ok=True)
        except Exception as e:
            messagebox.showerror("Error", f"Could not read workbook:\n{e}")
            self.set_status("Failed to read workbook.", ok=False)

    def load_headers(self):
        if not self.file_path or not self.sheet_var.get():
            messagebox.showwarning("Missing info", "Please select a file and sheet first.")
            return
        try:
            header_row = int(self.header_row_var.get())
            if header_row < 1:
                raise ValueError
        except ValueError:
            messagebox.showerror("Invalid input", "Header row must be a positive integer.")
            return

        try:
            df = pd.read_excel(
                self.file_path,
                sheet_name=self.sheet_var.get(),
                header=header_row - 1,
                nrows=0,
                engine="openpyxl",
            )
            self.headers = list(df.columns)
            self.header_vars = {h: tk.BooleanVar(value=True) for h in self.headers}

            for w in list(self.col_list.children.values()):
                w.destroy()
            for col in self.headers:
                ttk.Checkbutton(self.col_list, text=str(col), variable=self.header_vars[col]).pack(anchor="w", pady=1)

            self.set_status("Headers loaded. Select columns to preview/export.", ok=True)
        except Exception as e:
            messagebox.showerror("Error", f"Could not load headers:\n{e}")
            self.set_status("Failed to load headers.", ok=False)

    def preview_data(self):
        if not self._validate_preconditions():
            return

        def task():
            t0 = time.time()
            try:
                n = max(1, int(self.preview_rows_var.get()))
                header_row = int(self.header_row_var.get()) - 1
                selected = self.get_selected_columns()
                df = pd.read_excel(
                    self.file_path,
                    sheet_name=self.sheet_var.get(),
                    header=header_row,
                    usecols=selected,
                    nrows=n,
                    engine="openpyxl",
                )
                self._ui(lambda: self.render_preview(df))
                self._ui(lambda: self.set_status(f"Preview loaded ({len(df)} rows) in {time.time()-t0:.2f}s.", True))
            except Exception as e:
                self._ui(lambda: messagebox.showerror("Error", f"Could not load preview:\n{e}"))
                self._ui(lambda: self.set_status("Preview failed.", False))
            finally:
                self._ui(lambda: self._progress_done())

        self._start_worker(task, indeterminate=True)

    def export_csv_fast(self):
        if not self._validate_preconditions():
            return

        base = os.path.splitext(os.path.basename(self.file_path))[0]
        default_name = f"{to_snake_case(base)}_{to_snake_case(self.sheet_var.get())}.csv"
        out_path = filedialog.asksaveasfilename(
            title="Save CSV As",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if not out_path:
            return

        opt_snake = self.opt_snake.get()
        opt_dedup = self.opt_dedup.get()
        selected_cols = self.get_selected_columns()
        if not selected_cols:
            messagebox.showwarning("No columns", "Select at least one column.")
            return

        def task():
            t0 = time.time()
            try:
                wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
                ws = wb[self.sheet_var.get()]

                header_row_idx = int(self.header_row_var.get())
                if header_row_idx < 1:
                    raise ValueError("Header row must be >= 1")
                if ws.max_row is None or ws.max_row < header_row_idx:
                    raise ValueError("Header row exceeds total rows in sheet.")
                header_cells = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
                header_map = {str(h): i for i, h in enumerate(header_cells)}
                missing = [c for c in selected_cols if c not in header_map]
                if missing:
                    raise ValueError(f"Selected columns not found in header: {missing}")

                selected_idx = [header_map[c] for c in selected_cols]

                with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)

                    out_header = [to_snake_case(c) if opt_snake else c for c in selected_cols]
                    writer.writerow(out_header)

                    seen: set = set() if opt_dedup else set()

                    total_rows = ws.max_row - header_row_idx if ws.max_row and ws.max_row > header_row_idx else 0
                    self._ui(lambda: self._progress_reset(total_rows))

                    processed = 0
                    written = 0

                    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                        if self._cancel.is_cancelled():
                            self._ui(lambda: self.set_status("Export cancelled by user.", False))
                            return
                        out = []
                        for i in selected_idx:
                            v = row[i] if i < len(row) else None
                            out.append("" if v is None else v)

                        if opt_dedup:
                            key = tuple(out)
                            if key in seen:
                                processed += 1
                                if total_rows:
                                    self._ui(lambda: self._progress_step(1))
                                continue
                            seen.add(key)

                        writer.writerow(out)
                        processed += 1
                        written += 1
                        if total_rows:
                            if processed % 100 == 0 or processed == total_rows:
                                self._ui(lambda: self._progress_set(min(processed, total_rows)))

                    self._ui(lambda: self.set_status(
                        f"Exported {written} rows in {time.time()-t0:.2f}s → {os.path.basename(out_path)}", True
                    ))
                    self._ui(lambda: messagebox.showinfo("Done", f"CSV saved to:\n{out_path}"))
            except Exception as e:
                self._ui(lambda: messagebox.showerror("Error", f"Export failed:\n{e}"))
                self._ui(lambda: self.set_status("Export failed.", False))
            finally:
                self._ui(lambda: self._progress_done())

        self._start_worker(task, indeterminate=False)

    # ----------------------------- Worker/Progress -----------------------------
    def _start_worker(self, target, indeterminate: bool):
        if self._worker_thread and self._worker_thread.is_alive():
            messagebox.showwarning("Busy", "Please wait for the current task to finish or cancel it.")
            return
        self._cancel = CancelToken()
        self._progress_reset(0 if not indeterminate else None)
        if indeterminate:
            self.prog.configure(mode="indeterminate")
            self.prog.start(10)
        else:
            self.prog.configure(mode="determinate")
        self.cancel_btn.configure(state="normal")

        self._worker_thread = threading.Thread(target=target, daemon=True)
        self._worker_thread.start()

    def cancel_current(self):
        self._cancel.cancel()

    def _progress_reset(self, maximum: Optional[int]):
        if maximum is None:
            return
        self.prog.configure(mode="determinate", maximum=max(1, int(maximum)))
        self.prog["value"] = 0

    def _progress_set(self, value: int):
        try:
            self.prog["value"] = value
        except Exception:
            pass

    def _progress_step(self, step: int):
        try:
            self.prog.step(step)
        except Exception:
            pass

    def _progress_done(self):
        try:
            if str(self.prog["mode"]) == "indeterminate":
                self.prog.stop()
            self.prog["value"] = self.prog["maximum"] if self.prog["mode"] == "determinate" else 0
            self.cancel_btn.configure(state="disabled")
        except Exception:
            pass

    # ----------------------------- UI Helpers -----------------------------
    def _ui(self, fn):
        self.root.after(0, fn)

    def _validate_preconditions(self) -> bool:
        if not self.file_path or not self.sheet_var.get():
            messagebox.showwarning("Missing info", "Please select a file and sheet first.")
            return False
        if not self.headers:
            messagebox.showwarning("No headers", "Click 'Load Headers' first.")
            return False
        if not self.get_selected_columns():
            messagebox.showwarning("No columns", "Select at least one column.")
            return False
        return True


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelToCSVApp(root)
    root.mainloop()
